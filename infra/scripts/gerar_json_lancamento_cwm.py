import json
from openpyxl import load_workbook


def is_linha_vazia(row):
    return all(cell is None or str(cell).strip() == "" for cell in row)


def normalizar(valor):
    if valor is None:
        return ""
    return str(valor).strip().lower()


def to_str(valor):
    if valor is None:
        return ""
    return str(valor)


def processar_grupo(linhas_grupo):
    """
    Processa um único bloco (empresa)
    """

    if len(linhas_grupo) < 4:
        return None

    # -------------------------
    # 1. Nome da empresa
    # -------------------------
    nome_empresa = to_str(linhas_grupo[0][0])

    # -------------------------
    # 2. Cabeçalho
    # -------------------------
    headers = [normalizar(h) for h in linhas_grupo[1]]

    # -------------------------
    # 3. Saldo anterior
    # -------------------------
    saldo_anterior = ""
    linha_saldo = linhas_grupo[2]

    for idx, col in enumerate(headers):
        if col == "saldo":
            saldo_anterior = to_str(linha_saldo[idx])
            break

    # -------------------------
    # 4. Última linha = totais
    # -------------------------
    linha_total = linhas_grupo[-1]

    total_debito = ""
    total_credito = ""

    for idx, col in enumerate(headers):
        if col == "debito":
            total_debito = to_str(linha_total[idx])
        elif col == "credito":
            total_credito = to_str(linha_total[idx])

    # -------------------------
    # 5. Lançamentos
    # (da linha 4 até a penúltima)
    # -------------------------
    lancamentos = []

    for linha in linhas_grupo[3:-1]:
        if is_linha_vazia(linha):
            continue

        registro = {}

        for idx, col in enumerate(headers):
            if col:
                registro[col] = to_str(linha[idx])

        lancamentos.append(registro)

    return {
        "nome_empresa": nome_empresa,
        "saldo_anterior": saldo_anterior,
        "total_debito": total_debito,
        "total_credito": total_credito,
        "lancamentos": lancamentos
    }


def processar_planilha(caminho_arquivo):
    wb = load_workbook(caminho_arquivo, data_only=True)
    ws = wb.active

    grupos = []
    grupo_atual = []

    for row in ws.iter_rows(values_only=True):

        if is_linha_vazia(row):
            # fecha grupo atual
            if grupo_atual:
                grupos.append(grupo_atual)
                grupo_atual = []
        else:
            grupo_atual.append(row)

    # último grupo (caso não termine com linha vazia)
    if grupo_atual:
        grupos.append(grupo_atual)

    # -------------------------
    # Processa e printa cada JSON
    # -------------------------
    for grupo in grupos:
        payload = processar_grupo(grupo)

        if payload:
            print(json.dumps(payload))
            print("\n"*2)


if __name__ == "__main__":
    # processar_planilha("/home/fcervan/Documentos/doc-ai/conciliacao/nao-conciliado.xlsx")
    processar_planilha("/home/fcervan/Documentos/doc-ai/conciliacao/nao-conciliado-3-grupos.xlsx")
