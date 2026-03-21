from .base_parser import BaseParser
import json
from openpyxl import load_workbook
from io import BytesIO


class CwmParser(BaseParser):

    def parse(self, arquivo):

        lista_payload = []

        # 🔥 suporta bytes (S3) ou caminho local
        if isinstance(arquivo, bytes):
            arquivo = BytesIO(arquivo)

        wb = load_workbook(arquivo, data_only=True)
        ws = wb.active

        grupos = []
        grupo_atual = []

        for row in ws.iter_rows(values_only=True):

            if self.is_linha_vazia(row):
                # fecha grupo atual
                if grupo_atual:
                    grupos.append(grupo_atual)
                    grupo_atual = []
            else:
                grupo_atual.append(row)

        # último grupo (caso não termine com linha vazia)
        if grupo_atual:
            grupos.append(grupo_atual)

        # -------------------------
        # Processa e monta JSON
        # -------------------------
        for grupo in grupos:
            payload = self.processar_grupo(grupo)

            if payload:
                lista_payload.append(json.dumps(payload, ensure_ascii=False))

        return lista_payload

    def is_linha_vazia(self, row):
        return all(cell is None or str(cell).strip() == "" for cell in row)

    def normalizar(self, valor):
        if valor is None:
            return ""
        return str(valor).strip().lower()

    def to_str(self, valor):
        if valor is None:
            return ""
        return str(valor)

    def processar_grupo(self, linhas_grupo):

        if len(linhas_grupo) < 4:
            return None

        # -------------------------
        # 1. Nome da empresa
        # -------------------------
        nome_empresa = self.to_str(linhas_grupo[0][0])

        # -------------------------
        # 2. Cabeçalho
        # -------------------------
        headers = [self.normalizar(h) for h in linhas_grupo[1]]

        # -------------------------
        # 3. Saldo anterior
        # -------------------------
        saldo_anterior = ""
        linha_saldo = linhas_grupo[2]

        for idx, col in enumerate(headers):
            if col == "saldo":
                saldo_anterior = self.to_str(linha_saldo[idx])
                break

        # -------------------------
        # 4. Última linha = totais
        # -------------------------
        linha_total = linhas_grupo[-1]

        total_debito = ""
        total_credito = ""

        for idx, col in enumerate(headers):
            if col == "debito":
                total_debito = self.to_str(linha_total[idx])
            elif col == "credito":
                total_credito = self.to_str(linha_total[idx])

        # -------------------------
        # 5. Lançamentos
        # -------------------------
        lancamentos = []

        for linha in linhas_grupo[3:-1]:
            if self.is_linha_vazia(linha):
                continue

            registro = {}

            for idx, col in enumerate(headers):
                if col:
                    registro[col] = self.to_str(linha[idx])

            lancamentos.append(registro)

        return {
            "nome_empresa": nome_empresa,
            "saldo_anterior": saldo_anterior,
            "total_debito": total_debito,
            "total_credito": total_credito,
            "lancamentos": lancamentos
        }